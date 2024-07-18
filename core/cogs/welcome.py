from discord.ext import commands
import openpyxl.styles
from core.classes import Cog_Extension
import discord , os , shutil , openpyxl , random , requests

#用戶進出伺服器訊息

class welcome(Cog_Extension):

    class WelcomeMsgList():

        def __init__(self):
            self.msgChannel = None
            self.msgColor = []
            self.title = []
            self.msg = []
            self.thumbnail = []
            self.image = []

    def __init__(self , bot):
        super().__init__(bot)

    def loadWelcomeMsgList(self , guildID):
        if os.path.isfile(f"assets/welcomeMsg/{guildID}/welcomeMsg.xlsx"):
            welcomeMsgSheet = openpyxl.load_workbook(f"assets/welcomeMsg/{guildID}/welcomeMsg.xlsx").worksheets[0]
            welcomeMsgList = self.WelcomeMsgList()
            try:
                if welcomeMsgSheet.cell(row=2 , column=2).value:
                    welcomeMsgList.msgChannel = int(welcomeMsgSheet.cell(row=2 , column=2).value)
                else:
                    raise
            except:
                return None
            for i in range(2 , 8):
                totalRow = 0
                for j in range(2 , welcomeMsgSheet.max_row+1):
                    value = welcomeMsgSheet.cell(row=j , column=i).value
                    if value:
                        totalRow += 1
                    else:
                        break
                if welcomeMsgSheet.cell(row=2 , column=1).value == 'Y' and totalRow != 0:   #totalRow為0是沒有項目有設定
                    if i == 3:
                        for k in range(2 , totalRow+2):
                            try:
                                color = int(welcomeMsgSheet.cell(row=k , column=i).value , 16)
                                welcomeMsgList.msgColor.append(color)
                            except:
                                pass
                    elif i == 4:
                        for k in range(2 , totalRow+2):
                            welcomeMsgList.title.append(welcomeMsgSheet.cell(row=k , column=i).value)
                    elif i == 5:
                        for k in range(2 , totalRow+2):
                            welcomeMsgList.msg.append(welcomeMsgSheet.cell(row=k , column=i).value)
                    elif i == 6:
                        for k in range(2 , totalRow+2):
                            welcomeMsgList.image.append(welcomeMsgSheet.cell(row=k , column=i).value)
                    elif i == 7:
                        for k in range(2 , totalRow+2):
                            welcomeMsgList.thumbnail.append(welcomeMsgSheet.cell(row=k , column=i).value)
                else:
                    return None
            return welcomeMsgList
        else:
            return None

    @commands.Cog.listener()
    async def on_member_join(self , member):
        welcomeMsgList = self.loadWelcomeMsgList(member.guild.id)
        if welcomeMsgList:
            welcomeMsg = []
            if not welcomeMsgList.msgColor:
                welcomeMsg.append(0xffffff)
            else:
                welcomeMsg.append(welcomeMsgList.msgColor[random.randrange(len(welcomeMsgList.msgColor))])
            welcomeMsg.append(welcomeMsgList.title[random.randrange(len(welcomeMsgList.title))])
            welcomeMsg.append(welcomeMsgList.msg[random.randrange(len(welcomeMsgList.msg))])
            welcomeMsg.append(welcomeMsgList.image[random.randrange(len(welcomeMsgList.image))])
            welcomeMsg.append(welcomeMsgList.thumbnail[random.randrange(len(welcomeMsgList.thumbnail))])
            embed = discord.Embed(title=welcomeMsg[1] , description=welcomeMsg[2].replace("<user>" , f"<@{member.id}>") , color=welcomeMsg[0])
            embed.set_author(name=self.bot.user.name , icon_url=self.bot.user.avatar.url)
            embed.set_image(url=welcomeMsg[3])
            embed.set_thumbnail(url=welcomeMsg[4])
            channel = self.bot.get_channel(welcomeMsgList.msgChannel)
            await channel.send(embed=embed)

    @commands.Cog.listener()
    async def on_member_remove(self , member):
        if member.guild.id == 812669323020730398:
            channel = self.bot.get_channel(812669323020730401)
            await channel.send(f"<@{member.id}> 爆氣中離了")

    @discord.app_commands.command(name="sendwelcomemsglist" , description="傳送新人加入歡迎訊息列表檔案")
    async def sendwelcomemsglist(self , interaction:discord.Interaction):
        await interaction.response.defer()
        if not os.path.isfile(f"assets/welcomeMsg/{interaction.guild.id}/welcomeMsg.xlsx"):
            os.makedirs(f"assets/welcomeMsg/{interaction.guild.id}/temp" , exist_ok=True)
            book = openpyxl.Workbook()
            try:
                sheet = book.active
                sheet.append(("是否啟用歡迎訊息功能(Y/N)" , "歡迎訊息傳送頻道ID(注意下方要顯示完整數字才算設置成功，可在數字前加「'」)" , "訊息邊條顏色(可放多項隨機選擇)" , "標題(可放多項隨機選擇)" , "內容(可放多項隨機選擇)" , "主圖網址(可放多項隨機選擇)" , "縮圖網址(可放多項隨機選擇)"))
                sheet.append(("Y" , "1000000000000000000" , "FFFFFF" , "歡迎加入！" , "歡迎 <user> 加入！" , "https://media.printables.com/media/prints/744872/images/5815268_a3bff0c6-4b08-4082-83ff-e773b4aec2ad_f82d2ae9-4bba-4dc6-86a5-7f7b4f98e48a/thumbs/inside/1280x960/png/rick-roll-2.webp" , "https://media.printables.com/media/prints/744872/images/5815268_a3bff0c6-4b08-4082-83ff-e773b4aec2ad_f82d2ae9-4bba-4dc6-86a5-7f7b4f98e48a/thumbs/inside/1280x960/png/rick-roll-2.webp"))
                sheet.column_dimensions['A'].width = 30
                sheet.column_dimensions['B'].width = 80
                sheet.column_dimensions['C'].width = 35
                sheet.column_dimensions['D'].width = 50
                sheet.column_dimensions['E'].width = 50
                sheet.column_dimensions['F'].width = 50
                sheet.column_dimensions['G'].width = 50
                for i in range(1 , 8):
                    sheet.cell(row=1, column=i).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thick" , color="000000"))
                book.save(f"assets/welcomeMsg/{interaction.guild.id}/temp/welcomeMsg.xlsx")
                await interaction.followup.send(file=discord.File(f"assets/welcomeMsg/{interaction.guild.id}/temp/welcomeMsg.xlsx"))
                shutil.rmtree(f"assets/welcomeMsg/{interaction.guild.id}/temp" , ignore_errors=True)
            except Exception as e:
                print(e)
        else:
            await interaction.followup.send(file=discord.File(f"assets/welcomeMsg/{interaction.guild.id}/welcomeMsg.xlsx"))

    @discord.app_commands.command(name="editwelcomemsglist" , description="修改新人加入歡迎訊息（請附上歡迎訊息列表檔案）")
    @discord.app_commands.describe(file="歡迎訊息列表檔案")
    async def editwelcomemsglist(self , interaction:discord.Interaction , file:discord.Attachment):
        await interaction.response.defer()
        if file.filename == "welcomeMsg.xlsx":
            try:
                os.makedirs(f"assets/welcomeMsg/{interaction.guild.id}/temp" , exist_ok=True)       #先下載再取代，比較安全
                r=requests.get(file.url , stream=True)
                with open(f"assets/welcomeMsg/{interaction.guild.id}/temp/welcomeMsg.xlsx", "wb") as outFile:
                    shutil.copyfileobj(r.raw, outFile)
                shutil.copyfile(f"assets/welcomeMsg/{interaction.guild.id}/temp/welcomeMsg.xlsx" , f"assets/welcomeMsg/{interaction.guild.id}/welcomeMsg.xlsx")
                shutil.rmtree(f"assets/welcomeMsg/{interaction.guild.id}/temp" , ignore_errors=True)
            except OSError:
                await interaction.followup.send("寫入檔案時發生錯誤")
            except:
                await interaction.followup.send(f"請附上修改自「/sendwelcomemsglist」指令提供的檔案")
            else:
                await interaction.followup.send("歡迎訊息更新完成")
        else:
            await interaction.followup.send(f"請附上修改自「/sendwelcomemsglist」指令提供的檔案")

async def setup(bot):
    await bot.add_cog(welcome(bot))