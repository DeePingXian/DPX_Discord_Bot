#底層相關
import discord
from discord.ext import tasks
from discord.ext.commands import Bot
import shutil
import os
import json
import openpyxl
import requests
import time
import uuid

with open('settings.json' , 'r' , encoding = 'utf8') as json_file:
    json_data = json.load(json_file)

intents = discord.Intents.all()
bot=Bot(command_prefix=json_data['command_prefix'] , intents=intents)
bot.remove_command("help")

#載入cogs

CogFileList = []
for CogFile in os.listdir('core\\cogs'):
    if CogFile.endswith('.py'):
        bot.load_extension(f'core.cogs.{CogFile[:-3]}')
        CogFileList.append(CogFile)
        print(f"已載入{CogFile}")

#啟動時清除音樂機器人資料

DB = bot.get_cog('MySQL')
DB.Init()
if DB.test():
    print('MySQL 操作正常')
else:
    print('MySQL 操作錯誤')

#清空musicTemp

if os.path.isdir('assets/musicBot/musicTemp'):
    shutil.rmtree('assets/musicBot/musicTemp' , ignore_errors = True)

#載入答錄機訊息

AnsweringDict = {}
def loadAnsweringContent():
    AnsweringDict.clear()
    AnsweringSheet = openpyxl.load_workbook('assets\\answeringMachine\\answeringContent.xlsx').worksheets[0]
    for i in range(2 , AnsweringSheet.max_row+1):
        value = []
        for j in range(1 , 6):
            value.append(AnsweringSheet.cell(row=i, column=j).value)
        AnsweringDict[AnsweringSheet.cell(row=i, column=1).value] = value        
loadAnsweringContent()

@bot.command()
async def sendansweringcontentlist(ctx):
    await ctx.send(file=discord.File('assets\\answeringMachine\\answeringContent.xlsx'))

#修改答錄機訊息

@bot.command()
async def editansweringcontentlist(ctx):
    try:
        _ = ctx.message.attachments[0]
    except:
        await ctx.reply('請附上修改自「!!sendansweringcontentlist」指令提供的檔案')
    else:
        if ctx.message.attachments[0].url[0:26] == "https://cdn.discordapp.com" and ctx.message.attachments[0].url[ctx.message.attachments[0].url.rfind('/')+1:] == 'answeringContent.xlsx':
            try:
                if not os.path.isdir('assets/answeringMachine/temp'):         #先下載再取代，比較安全
                    os.mkdir('assets/answeringMachine/temp')
                r=requests.get(ctx.message.attachments[0].url , stream=True)
                with open('assets\\answeringMachine\\temp\\answeringContent.xlsx', 'wb') as outFile:
                    shutil.copyfileobj(r.raw, outFile)
                shutil.copyfile('assets\\answeringMachine\\temp\\answeringContent.xlsx' , 'assets\\answeringMachine\\answeringContent.xlsx')
                if os.path.isfile('assets\\answeringMachine\\temp\\answeringContent.xlsx'):
                    os.remove('assets\\answeringMachine\\temp\\answeringContent.xlsx')
            except:
                await ctx.reply('寫入檔案時發生錯誤')
            else:
                loadAnsweringContent()
                await ctx.send('應答機更新完成')
        else:
            await ctx.reply('請附上修改自「!!sendansweringcontentlist」指令提供的檔案')

#啟動訊息

@bot.event
async def on_ready():
    channel = bot.get_channel(json_data['log_channel_id'])
    game = discord.Game(f'輸入{json_data["command_prefix"]}help查詢指令')
    await bot.change_presence(status=discord.Status.online, activity=game)
    ping = round (bot.latency * 1000)
    print(f'啟動完成 與 Discord 延遲為 {ping} ms')
    await channel.send('啟動完成')
    DB.CleanAllMusicTable()
    autoLog.start()

@bot.event
async def on_message(message):

    #訊息log
    
    if message.content != '' or message.attachments != []:
        DB.PutMessageLog(message , None , None , 0)

    #應答機

    if message.author != bot.user:
        for i in AnsweringDict:
            if ( message.content.lower() == i.lower() and ( AnsweringDict[i][3] == 'Y' or message.content == i ) ) or ( ( i in message.content or (i.lower() in message.content.lower() and AnsweringDict[i][3] == 'Y' ) ) and AnsweringDict[i][4] == 'Y' ):
                if AnsweringDict[i][1] != None:
                    await message.channel.send(AnsweringDict[i][1])
                if AnsweringDict[i][2] != None:
                    try:
                        await message.channel.send(file=discord.File(f'assets\\answeringMachine\\{AnsweringDict[i][2]}'))
                    except:
                        pass

    await bot.process_commands(message)

#取得本 Discord 伺服器的 ID

@bot.command()
async def getguildid(ctx):
    await ctx.send(f'本伺服器ID：{ctx.guild.id}')

#更新資料庫網頁存取token

@tasks.loop(seconds=60)
async def waitForUpdateToken():
    nowHour = time.gmtime().tm_hour
    nowMinute = time.gmtime().tm_min
    if nowHour == 0 and nowMinute >= 0 and nowMinute <= 1:
        updateToken.start()
        waitForUpdateToken.stop()

@tasks.loop(hours=24)
async def updateToken():
    DB.UpdateToken(str(uuid.uuid4())[:8])

#取得存取網頁之token

@bot.command()
async def getwebtoken(ctx):
    await ctx.send(f'目前存取網頁之 token：\n{DB.GetToken()}\ntoken 每天 UTC 0:00／CST 8:00 更新')

#bot狀態回報

@bot.command()
async def status(ctx):
    if ctx.author.id == json_data['adminID']:
        DBStatus = ''
        if DB.test():
            DBStatus = '正常'
        else:
            DBStatus = '錯誤'
        ping = round (bot.latency * 1000)
        CogFileString = ""
        for CogFile in CogFileList:
            CogFileString += f'{CogFile}\n'        
        embed = discord.Embed(title="bot 狀態一覽", description='\u200b' , color=0xc0c0c0)
        embed.set_author(name="DPX discord bot" , url=discord.Embed.Empty , icon_url=json_data['BotIconUrl'])
        embed.add_field(name="類別", value='與 Discord 網路延遲\nMySQL 操作狀態' , inline=True)
        embed.add_field(name="狀態", value=f'{ping} ms\n{DBStatus}' , inline=True)
        embed.add_field(name='\u200b', value='\u200b' , inline=False)
        embed.add_field(name="已載入之cog", value=CogFileString , inline=False)
        await ctx.send(embed=embed)
    else:
        await ctx.reply('此功能僅限開此bot的管理員使用')

@tasks.loop(hours=1)
async def autoLog():
    await log()

async def log():
    channel = bot.get_channel(json_data['log_channel_id'])
    word = f'與 Discord 延遲為 {round(bot.latency * 1000)} ms'
    if DB.test():
        word += '\nMySQL 操作正常'
    else:
        word += '\nMySQL 操作錯誤'
    await channel.send(word)

#錯誤訊息

@bot.event
async def on_error(event, *args, **kwargs):
    message = args[0] #Gets the message object
    await SendErrorMessage(message)

@bot.event
async def on_command_error(ctx, error):
    await SendErrorMessage(ctx.message)

async def SendErrorMessage(message):
    await message.reply('發生錯誤')
    await message.channel.send(file=discord.File(json_data['error_message_file']))

if __name__ == "__main__":
    bot.run(json_data['TOKEN'])